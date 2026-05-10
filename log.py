from openpyxl import load_workbook

def write(training_time, final_loss, parameters, config, loss_diff):
    wb = load_workbook('ML Stats.xlsx')
    sheet = wb.active

    first_empty_row = sheet.max_row + 1
    
    sheet[f'A{first_empty_row}'] = f'Model {first_empty_row - 1}'
    sheet[f'B{first_empty_row}'] = training_time
    sheet[f'C{first_empty_row}'] = final_loss
    sheet[f'D{first_empty_row}'] = parameters
    for i, (key, value) in enumerate(config.items(), start=5):
        sheet[f'{chr(64 + i)}{first_empty_row}'] = f'{value}'
    sheet[f'{chr(64 + i + 1)}{first_empty_row}'] = f'{loss_diff}'
    wb.save('ML Stats.xlsx')
    print(f"Logged results for Model {first_empty_row - 1} to Excel.")